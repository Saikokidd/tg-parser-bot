"""
Генерация .xlsx файла с выгрузкой лидов и их родственников.

Структура:
- Лист "Родственники" — все поля + гиперссылка на военного на втором листе
- Лист "Военные" — все поля

Используется openpyxl.
"""
import io
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bot.parser.military_parser import status_label

logger = logging.getLogger(__name__)


# ──────────── Стили ────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
LINK_FONT = Font(color="0563C1", underline="single")

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


# ──────────── Колонки ────────────
# (заголовок, ширина)

MILITARY_COLUMNS = [
    ("ID", 6),
    ("ФИО", 30),
    ("ДР", 12),
    ("Статус", 14),
    ("Б/Ч", 12),
    ("Позывной", 14),
    ("Источник", 30),
    ("Доп инфа", 30),
]

RELATIVE_COLUMNS = [
    ("ID", 6),
    ("ФИО", 30),
    ("ДР", 12),
    ("Адрес", 40),
    ("Телефоны", 40),
    ("Регион", 24),
    ("СНИЛС", 16),
    ("ИНН", 16),
    ("Паспорт", 14),
    ("Почта", 24),
    ("Закреплён за", 30),
    ("Доп поля", 30),
]


# ──────────── Помощники форматирования ────────────

STD_EXTRA_KEYS_RELATIVE = {"snils", "inn", "passport", "email", "emails", "operator", "region", "phones_all", "old_operator", "tz_offset"}
STD_EXTRA_KEYS_MILITARY = {"unit", "callsign", "note", "source"}


def _fmt_date(d):
    if not d:
        return ""
    if hasattr(d, "strftime"):
        return d.strftime("%d.%m.%Y")
    return str(d)


def _fmt_dt(dt):
    if not dt:
        return ""
    if hasattr(dt, "strftime"):
        return dt.strftime("%d.%m.%Y %H:%M")
    return str(dt)


# Маппинг hlr_status → emoji + текст
HLR_STATUS_LABELS = {
    "available":         "✅",
    "unavailable":       "❌ выкл",
    "not_exists":        "⛔ нет",
    "in_work":           "⏳",
    "pending":           "⏳",
    "skipped_operator":  "⊝",
    "error":             "⚠️",
}


def _fmt_op(operator: str | None, old_operator: str | None) -> str:
    """Оператор с историей переноса: 'ВымпелКом→МегаФон' если номер перенесён."""
    op = operator or "—"
    if old_operator and old_operator != operator:
        return f"{old_operator}→{op}"
    return op


def _fmt_phones(phones: list[dict], legacy_phone: str | None,
                 legacy_operator: str | None,
                 phones_all: list[dict] | None = None,
                 legacy_old_operator: str | None = None) -> str:
    """
    Формирует многострочную ячейку с телефонами.
    Оператор с историей переноса (old_operator), если номер был перенесён.
    Primary номер всегда первым.
    """
    if phones:
        lines = []
        for i, p in enumerate(phones):
            phone = p.get("phone") or ""
            op = _fmt_op(p.get("operator"), p.get("old_operator"))
            status_key = p.get("hlr_status") or ""
            label = HLR_STATUS_LABELS.get(status_key, "")
            line = f"{phone} ({op})"
            if label:
                line = f"{line} {label}"
            if i == 0 and line.startswith(("+", "-", "=")):
                line = "'" + line
            lines.append(line)
        return "\n".join(lines)

    # phones_all — сауроновский dp-формат: все номера с операторами стопкой.
    if phones_all:
        lines = []
        for i, p in enumerate(phones_all):
            ph = p.get("phone") or ""
            op = _fmt_op(p.get("operator"), p.get("old_operator"))
            line = f"{ph} ({op})"
            if i == 0 and line.startswith(("+", "-", "=")):
                line = "'" + line
            lines.append(line)
        return "\n".join(lines)

    # Fallback на legacy
    if legacy_phone:
        parts = [legacy_phone]
        if legacy_operator or legacy_old_operator:
            parts.append(f"({_fmt_op(legacy_operator, legacy_old_operator)})")
        result = " ".join(parts)
        if result.startswith(("+", "-", "=")):
            result = "'" + result
        return result
    return ""


def _fmt_extra_custom(extra: dict, std_keys: set) -> str:
    """
    Вернуть кастомные поля extra в виде 'ключ: значение; ключ: значение'.
    Стандартные ключи (snils, inn и т.д.) пропускаются.
    """
    if not extra:
        return ""
    parts = []
    for k, v in extra.items():
        if k in std_keys or not v:
            continue
        parts.append(f"{k}: {v}")
    return "; ".join(parts)


def _set_header(ws, columns: list):
    """Заголовок листа: жирно, белым по синему, фиксированная высота"""
    for col_idx, (title, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"  # шапка зафиксирована


def _apply_cell_style(cell, is_link: bool = False):
    cell.alignment = CELL_ALIGN
    cell.border = THIN_BORDER
    if is_link:
        cell.font = LINK_FONT


# ──────────── Главная функция ────────────

def build_xlsx(military_records: list, relatives: list,
               manager_label: str = None) -> bytes:
    """
    Сгенерировать .xlsx файл (bytes).

    military_records — список военных (dict с full_name, birth_date, status, extra и т.д.)
    relatives — список родственников (dict с linked_military)
    manager_label — название (для имени файла, не используется внутри книги)
    """
    wb = Workbook()

    # ──────────── ЛИСТ 1: РОДСТВЕННИКИ ────────────
    ws_rel = wb.active
    ws_rel.title = "Родственники"
    _set_header(ws_rel, RELATIVE_COLUMNS)

    # Запоминаем строки военных на 2-м листе для гиперссылок
    # (заполним после построения 2-го листа, но текст уже подготовим)

    # Индекс военных по id → row на листе "Военные"
    # row=1 — заголовок, военные начнут с row=2
    military_row_by_id = {}
    for idx, m in enumerate(military_records, start=2):
        military_row_by_id[m["id"]] = idx

    # Заполняем родственников
    for r_idx, r in enumerate(relatives, start=2):
        extra = r.get("extra") or {}
        custom_fields = _fmt_extra_custom(extra, STD_EXTRA_KEYS_RELATIVE)

        # Поле "Закреплён за" — может быть несколько военных
        linked = r.get("linked_military") or []
        # Сортируем: сначала те что есть в нашей выгрузке (с military_row_by_id)
        linked_in_export = [m for m in linked if m["id"] in military_row_by_id]
        linked_outside = [m for m in linked if m["id"] not in military_row_by_id]

        # Текст ячейки — все военные через перенос строки
        linked_lines = []
        for m in linked_in_export + linked_outside:
            birth = m.get("birth_date")
            if birth:
                # birth приходит строкой 'YYYY-MM-DD'
                try:
                    birth = datetime.strptime(birth, "%Y-%m-%d").strftime("%d.%m.%Y")
                except Exception:
                    pass
            linked_lines.append(f"{m['full_name']} ({birth or '—'})")
        linked_text = "\n".join(linked_lines) if linked_lines else ""

        # Многострочная ячейка с номерами + статусами HLR
        phones_text = _fmt_phones(
            r.get("phones") or [],
            r.get("phone"),
            extra.get("operator"),
            extra.get("phones_all"),
            extra.get("old_operator"),
        )
        row = [
            r.get("id"),
            r.get("full_name") or "",
            _fmt_date(r.get("birth_date")),
            r.get("address") or "",
            phones_text,
            (f"{extra.get('region')} ({extra.get('tz_offset')})"
             if extra.get("region") and extra.get("tz_offset")
             else (extra.get("region") or "")),
            extra.get("snils") or "",
            extra.get("inn") or "",
            extra.get("passport") or "",
            ", ".join(extra.get("emails") or []) or (extra.get("email") or ""),
            linked_text,
            custom_fields,
        ]

        for col_idx, value in enumerate(row, start=1):
            cell = ws_rel.cell(row=r_idx, column=col_idx, value=value)
            _apply_cell_style(cell)

        # Гиперссылка на ячейку "Закреплён за" — на ПЕРВОГО военного в списке (который в выгрузке)
        if linked_in_export:
            first = linked_in_export[0]
            target_row = military_row_by_id[first["id"]]
            link_cell = ws_rel.cell(row=r_idx, column=11)  # "Закреплён за"
            link_cell.hyperlink = f"#'Военные'!A{target_row}"
            _apply_cell_style(link_cell, is_link=True)

    # ──────────── ЛИСТ 2: ВОЕННЫЕ ────────────
    ws_mil = wb.create_sheet("Военные")
    _set_header(ws_mil, MILITARY_COLUMNS)

    for m_idx, m in enumerate(military_records, start=2):
        extra = m.get("extra") or {}

        row = [
            m.get("id"),
            m.get("full_name") or "",
            _fmt_date(m.get("birth_date")),
            status_label(m.get("status")),
            extra.get("unit") or "",
            extra.get("callsign") or "",
            m.get("source_name") or extra.get("source") or "",
            extra.get("note") or "",
        ]

        for col_idx, value in enumerate(row, start=1):
            cell = ws_mil.cell(row=m_idx, column=col_idx, value=value)
            _apply_cell_style(cell)

    # ──────────── Сохраняем в bytes ────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def make_filename(manager_label: str = None) -> str:
    """parser_export_<manager>_2026-05-08_14-30.xlsx"""
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if manager_label:
        safe = "".join(c if c.isalnum() else "_" for c in manager_label)
        return f"parser_export_{safe}_{ts}.xlsx"
    return f"parser_export_{ts}.xlsx"
