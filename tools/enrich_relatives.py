#!/usr/bin/env python3
"""
Шаг 3 (v2): обогащение военных batch sauron_links_20260720 родственниками
по приоритетам фонда.

ОТБОР на военного (потолок 2 закрепления, добивка до 2 КОНТАКТНЫХ):
  очередь приоритета: Мать -> Отец -> Супруга -> ребёнок(макс 1) -> Брат/Сестра -> прочее
  идём по очереди, пробиваем родственника (probit_person, счёт pvl);
  если у него есть телефон — закрепляем; если нет — берём следующего,
  пока не наберём 2 контактных или очередь не кончится.

Пропускаем военных, у которых у Сони уже есть родственники (идемпотентность).

    venv/bin/python -m tools.enrich_relatives /root/exports/svyazi_est_2186.xlsx --limit 20            # DRY-RUN
    venv/bin/python -m tools.enrich_relatives /root/exports/svyazi_est_2186.xlsx --limit 20 --commit   # пилот
    venv/bin/python -m tools.enrich_relatives /root/exports/svyazi_est_2186.xlsx --commit               # всё
"""
import argparse
import asyncio
import re
from datetime import date

from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv()

from bot.db.connection import get_pool
from bot.db.queries import insert_relative_v2, link_military_relative
from bot.services.probiv_service import probit_person
from bot.handlers.probiv import _build_relative_data_from_template

SONYA = 111
BATCH = "sauron_links_20260720"
OFFICE = "pvl"
CAP = 2


def norm(s: str) -> str:
    s = (s or "").strip().lower().replace("ё", "е").rstrip(" .,;")
    return re.sub(r"\s+", " ", s)


def parse_dr(dr: str):
    try:
        d, m, y = dr.split(".")
        return date(int(y), int(m), int(d))
    except Exception:
        return None


def relation_rank(rel: str) -> int:
    r = (rel or "").strip().lower()
    if "мать" in r:
        return 0
    if "отец" in r:
        return 1
    if "супруг" in r:
        return 2
    if "сын" in r or "дочь" in r:
        return 3
    if "брат" in r or "сестра" in r:
        return 4
    return 5


def read_file(path: str):
    ws = load_workbook(path, read_only=True).active
    data = {}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        subj_fio = str(r[1]).strip() if r[1] else ""
        subj_dr = str(r[2]).strip() if r[2] else ""
        rel_fio = str(r[3]).strip() if r[3] else ""
        rel_dr = str(r[4]).strip() if r[4] else ""
        relation = str(r[6]).strip() if r[6] else ""
        if not subj_fio or not subj_dr:
            continue
        if not rel_fio or rel_fio in ("—", "") or relation.lower() == "нет связей":
            continue
        key = (norm(subj_fio), subj_dr)
        data.setdefault(key, []).append((rel_fio, rel_dr, relation))
    for key in data:
        data[key].sort(key=lambda t: relation_rank(t[2]))
    return data


def build_queue(rels):
    queue, child_used = [], False
    for rf, rd, rel in rels:
        if relation_rank(rel) == 3:
            if child_used:
                continue
            child_used = True
        queue.append((rf, rd, rel))
    return queue


async def main(path, limit, commit):
    file_map = read_file(path)
    pool = await get_pool()

    async with pool.acquire() as conn:
        mils = await conn.fetch(
            "SELECT id, full_name, birth_date FROM persons_military "
            "WHERE added_by = $1 AND extra->'moved'->>'batch' = $2 ORDER BY id",
            SONYA, BATCH,
        )
        rows = await conn.fetch(
            "SELECT DISTINCT mr.military_id FROM military_relatives mr "
            "JOIN persons_military pm ON pm.id = mr.military_id "
            "WHERE pm.added_by = $1 AND pm.extra->'moved'->>'batch' = $2",
            SONYA, BATCH,
        )
        already = {r["military_id"] for r in rows}

    tasks = []
    for m in mils:
        if m["id"] in already:
            continue
        key = (norm(m["full_name"]), m["birth_date"].strftime("%d.%m.%Y"))
        rels = file_map.get(key)
        if rels:
            tasks.append((m["id"], m["full_name"], build_queue(rels)))

    if limit:
        tasks = tasks[:limit]

    print(f"Военных в batch: {len(mils)}, уже с роднёй (пропуск): {len(already)}")
    print(f"К обогащению: {len(tasks)} (потолок {CAP} контактных)")
    if limit:
        print(f"(ПИЛОТ: --limit {limit})")

    if not commit:
        print("\n[DRY-RUN] Первые 5 военных и очередь отбора:")
        for mil_id, fio, queue in tasks[:5]:
            print(f"  {mil_id} {fio}:")
            for rf, rd, rel in queue:
                print(f"      {rel:24} {rf} {rd}")
        print("\nДля внесения: --commit")
        return

    stat = dict(mil=0, probed=0, linked=0, no_phone=0, errors=0, dup=0)
    for mil_id, fio, queue in tasks:
        stat["mil"] += 1
        got = 0
        seen_rel = set()          # (norm ФИО, ДР) уже закреплённых у этого военного
        for rf, rd, rel in queue:
            if got >= CAP:
                break
            rel_key = (norm(rf), rd)
            if rel_key in seen_rel:   # тот же человек уже взят — пропускаем дубль
                stat["dup"] += 1
                continue
            stat["probed"] += 1
            try:
                res = await probit_person(rf, parse_dr(rd), manager_id=SONYA,
                                          context="tool", military_id=mil_id, office=OFFICE)
            except Exception as e:
                stat["errors"] += 1
                print(f"  ⚠️ пробив {rf}: {e}")
                continue
            tmpl = getattr(res, "relative_template", None)
            data = await _build_relative_data_from_template(tmpl) if tmpl else None
            if not data or not data.get("phone"):
                stat["no_phone"] += 1
                continue
            data.setdefault("extra", {})["relation"] = rel
            try:
                rec = await insert_relative_v2(data, SONYA)
                if await link_military_relative(mil_id, rec["id"], SONYA):
                    stat["linked"] += 1
                    got += 1
                    seen_rel.add(rel_key)
            except Exception as e:
                stat["errors"] += 1
                print(f"  ⚠️ вставка {rf}: {e}")
        print(f"  [{stat['mil']}/{len(tasks)}] военный {mil_id}: закреплено {got}")

    print(f"\n✅ Военных обработано: {stat['mil']}")
    print(f"   пробивов: {stat['probed']}, закреплено: {stat['linked']}")
    print(f"   без телефона: {stat['no_phone']}, дублей пропущено: {stat['dup']}, ошибок: {stat['errors']}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--commit", action="store_true")
    args = ap.parse_args()
    asyncio.run(main(args.path, args.limit, args.commit))