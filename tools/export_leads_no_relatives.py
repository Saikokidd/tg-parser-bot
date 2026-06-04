"""
Разовая утилита: выгрузить N лидов с ДР и без родственников в xlsx.

Отбор:
- birth_date IS NOT NULL       — только с датой рождения
- exported_at IS NULL          — ещё не выгружали
- нет ни одной записи в military_relatives
- ORDER BY created_at DESC     — самые свежие первыми
- LIMIT 500 (настраивается через --limit)

После успешной выгрузки помечает все взятые лиды как exported_at = NOW().
Если файл не записался — UPDATE откатывается, лиды остаются "свежими".

Использование:
    cd ~/projects/tg-parser-bot
    venv/bin/python -m tools.export_leads_no_relatives
    venv/bin/python -m tools.export_leads_no_relatives --limit 1000
    venv/bin/python -m tools.export_leads_no_relatives --dry-run     # без UPDATE
"""
import asyncio
import sys
import logging
import argparse
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from bot.db.connection import get_pool, close_pool

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("export_leads_no_relatives")

OUTPUT_DIR = Path("/root/exports")
DEFAULT_LIMIT = 500


async def fetch_leads(limit: int) -> list[dict]:
    """
    Выборка лидов:
    - С ДР
    - Без родственников
    - Не выгружали ранее
    - Самые свежие
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT pm.id, pm.full_name, pm.birth_date, pm.extra
            FROM persons_military pm
            WHERE pm.birth_date IS NOT NULL
              AND pm.exported_at IS NULL
              AND NOT EXISTS (
                  SELECT 1 FROM military_relatives mr WHERE mr.military_id = pm.id
              )
            ORDER BY pm.created_at DESC
            LIMIT $1
            """,
            limit,
        )
        return [dict(r) for r in rows]


def build_xlsx(leads: list[dict], output_path: Path) -> None:
    """
    Создаёт xlsx с 3 колонками: ФИО / ДР / Доп.инфа.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Лиды без родственников"

    # Заголовок
    headers = ["ФИО", "ДР", "Доп. инфа"]
    widths = [40, 12, 50]
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Данные
    for row_idx, lead in enumerate(leads, start=2):
        ws.cell(row=row_idx, column=1, value=lead["full_name"] or "")

        birth = lead.get("birth_date")
        ws.cell(row=row_idx, column=2,
                value=birth.strftime("%d.%m.%Y") if birth else "")

        extra = lead.get("extra") or {}
        note = extra.get("note") or ""
        ws.cell(row=row_idx, column=3, value=note)

    # Замораживаем шапку
    ws.freeze_panes = "A2"
    wb.save(output_path)


async def mark_exported(ids: list[int]) -> None:
    """Пометить выгруженные лиды exported_at = NOW()"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE persons_military SET exported_at = NOW() WHERE id = ANY($1::int[])",
            ids,
        )


async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=DEFAULT_LIMIT,
                        help=f"Сколько лидов выгружать (default {DEFAULT_LIMIT})")
    parser.add_argument("--dry-run", action="store_true",
                        help="Только сформировать файл, не помечать как выгруженные")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    logger.info(f"Выборка до {args.limit} лидов...")
    leads = await fetch_leads(args.limit)
    actual_count = len(leads)
    logger.info(f"Найдено: {actual_count} лидов")

    if actual_count == 0:
        logger.warning("Подходящих лидов нет. Файл не создан.")
        await close_pool()
        return

    if actual_count < args.limit:
        logger.warning(
            f"Запрошено {args.limit}, доступно {actual_count}. "
            f"Выгружаем что есть."
        )

    # Имя файла с датой и фактическим количеством
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"leads_no_relatives_{actual_count}_{ts}.xlsx"

    logger.info(f"Запись в {output_path}...")
    build_xlsx(leads, output_path)
    logger.info(f"Файл записан: {output_path} ({output_path.stat().st_size // 1024} KB)")

    if args.dry_run:
        logger.info("DRY-RUN: лиды НЕ помечены как exported_at. Завершено.")
    else:
        ids = [lead["id"] for lead in leads]
        logger.info(f"Помечаю {len(ids)} лидов как exported_at = NOW()...")
        await mark_exported(ids)
        logger.info(f"Готово. Помечено {len(ids)} лидов.")

    await close_pool()


if __name__ == "__main__":
    asyncio.run(main())
